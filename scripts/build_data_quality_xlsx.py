"""Build DATA_QUALITY.xlsx in the repo root (run from repo root: python scripts/build_data_quality_xlsx.py)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "DATA_QUALITY.xlsx"

INTRO = (
    "Automated scan of the five CSVs in data/: structural issues (dup keys, orphan FKs, etc.) were clean in this extract. "
    "Blank close_date for non-Closed outcomes (Pending, Withdrawn, …) is expected business state, not a defect. "
    "Business-rule asset checks (Closed ⇒ close_date; both dates ⇒ days_to_close) are implemented in checks.py."
)

HEADERS = [
    "File name",
    "Field(s) — business meaning",
    "Issue (class)",
    "Observed in current raw files?",
    "How resolved",
    "Programmatic check",
    "Code location",
]

ROWS: list[tuple[str, ...]] = [
    (
        "transactions.csv",
        "transaction_id — Stable identifier for one M&A event (one row in the fact table).",
        "Missing / blank transaction_id",
        "No (0 rows)",
        "Rows moved to staging.transactions_quarantine (missing_transaction_id); excluded from staging.transactions",
        "trim(coalesce(transaction_id, '')) = '' in quarantine + main filter",
        "william_blair_de/assets/staging.py (stg_transactions)",
    ),
    (
        "transactions.csv",
        "deal_size_mm — Transaction enterprise value in millions of USD; should not be negative.",
        "Negative deal_size_mm",
        "No (0 rows)",
        "Quarantine reason negative_deal_size; excluded from clean staging.transactions",
        "TRY_CAST(deal_size_mm AS DOUBLE) < 0",
        "william_blair_de/assets/staging.py",
    ),
    (
        "transactions.csv",
        "acquirer_id — Buyer entity key; must match acquirer master so deals roll up to a real counterparty.",
        "acquirer_id not present in acquirer dimension",
        "No (0 orphan rows on scan)",
        "Not dropped in staging; surfaced as failed check",
        "NOT EXISTS (SELECT 1 FROM staging.acquirers a WHERE a.acquirer_id = t.acquirer_id)",
        "william_blair_de/assets/checks.py (stg_transactions_fk_acquirer)",
    ),
    (
        "transactions.csv",
        "target_id — Seller / portfolio company key; must match target master for sector and financial context.",
        "target_id not present in target dimension",
        "No",
        "Same as above",
        "NOT EXISTS (… staging.targets …)",
        "william_blair_de/assets/checks.py (stg_transactions_fk_target)",
    ),
    (
        "transactions.csv",
        "announce_date, close_date — When the deal was made public vs when it legally completed; close should not precede announce.",
        "close_date earlier than announce_date",
        "No",
        "Staging keeps row; check fails so issue is visible",
        "close_date < announce_date when both non-null",
        "william_blair_de/assets/checks.py (stg_transactions_date_order)",
    ),
    (
        "transactions.csv",
        "transaction_id — Must be unique so each deal is counted once in analytics.",
        "Duplicate transaction_id",
        "No",
        "Staging does not dedupe; check fails",
        "GROUP BY transaction_id HAVING COUNT(*) > 1",
        "william_blair_de/assets/checks.py (stg_transactions_unique_ids)",
    ),
    (
        "acquirers.csv",
        "acquirer_id — Primary key for the buyer (strategic company or PE sponsor).",
        "Blank / whitespace-only acquirer_id",
        "No",
        "Row dropped from staging.acquirers",
        "WHERE acquirer_id IS NOT NULL AND trim(acquirer_id) <> ''",
        "william_blair_de/assets/staging.py (stg_acquirers)",
    ),
    (
        "acquirers.csv",
        "acquirer_type — Strategic vs Financial Sponsor (PE); drives which metrics make sense.",
        "Inconsistent acquirer_type labels (casing / synonyms)",
        "Partially (data already clean; CASE still normalizes)",
        "Map known lowercase forms to canonical labels; else trimmed string",
        "CASE + trim in SQL",
        "william_blair_de/assets/staging.py (stg_acquirers)",
    ),
    (
        "acquirers.csv",
        "acquirer_id — One row per acquirer in the dimension.",
        "Duplicate acquirer_id",
        "No",
        "Staging does not dedupe; check fails",
        "GROUP BY acquirer_id HAVING COUNT(*) > 1 on staging.acquirers",
        "william_blair_de/assets/checks.py (stg_acquirers_unique_ids)",
    ),
    (
        "targets.csv",
        "target_id — Primary key for the company being acquired or pursued.",
        "Blank target_id",
        "No",
        "Row dropped from staging.targets",
        "WHERE target_id IS NOT NULL AND trim(target_id) <> ''",
        "william_blair_de/assets/staging.py (stg_targets)",
    ),
    (
        "targets.csv",
        "target_id — One row per target company.",
        "Duplicate target_id",
        "No",
        "Check fails",
        "GROUP BY target_id HAVING COUNT(*) > 1 on staging.targets",
        "william_blair_de/assets/checks.py (stg_targets_unique_ids)",
    ),
    (
        "acquirer_financials.csv",
        "acquirer_id — Links annual financials to strategic acquirer (PE should not appear per data dictionary).",
        "Blank acquirer_id",
        "No",
        "Row dropped from staging.acquirer_financials",
        "WHERE acquirer_id IS NOT NULL AND trim(acquirer_id) <> ''",
        "william_blair_de/assets/staging.py (stg_acquirer_financials)",
    ),
    (
        "acquirer_financials.csv",
        "revenue_mm, ebitda_mm, ebitda_margin_pct, net_debt_mm, cash_on_hand_mm, total_assets_mm, market_cap_mm, employee_count, fiscal_year — Annual snapshot for capacity / leverage context.",
        "Non-numeric numeric fields after ingest",
        "Not separately counted",
        "TRY_CAST → NULL for bad values (soft quarantine)",
        "Implicit TRY_CAST(...) on each numeric column",
        "william_blair_de/assets/staging.py (stg_acquirer_financials)",
    ),
    (
        "sector_multiples.csv",
        "median_ev_ebitda, p25/p75, median_ev_revenue, deal_count, fiscal_year, fiscal_quarter, sector — Market benchmark multiples for comps.",
        "Non-numeric medians / counts",
        "Not separately counted",
        "TRY_CAST → NULL; model joins may miss benchmark",
        "Implicit TRY_CAST",
        "william_blair_de/assets/staging.py (stg_sector_multiples)",
    ),
    (
        "All ingested CSVs",
        "Date columns (announce_date, close_date, founded_date) — Calendar semantics for timelines and fiscal alignment.",
        "Date strings using / instead of -",
        "Not observed (all ISO in sample)",
        "replace(trim(date_col), '/', '-') before date cast",
        "Same expression in staging SQL",
        "william_blair_de/assets/staging.py (all stg_* assets)",
    ),
    (
        "All ingested CSVs",
        "Business keys and labels (acquirer_id, target_id, transaction_id, names, sectors, text) — Join keys and reporting attributes.",
        "Leading/trailing spaces on business keys / text",
        "Not observed on IDs (scan)",
        "trim / upper on keys in staging",
        "Visual + staging predicates",
        "william_blair_de/assets/staging.py",
    ),
    (
        "transactions.csv",
        "outcome, close_date — Deal state vs legal completion date; only Closed deals should have a close.",
        "Expected: null close_date when outcome is not Closed (Pending, Withdrawn, Terminated, Rumored)",
        "Yes (197 rows in sample; 303 Closed all have close_date)",
        "No pipeline change — valid lifecycle; distinguish from Closed-without-close errors",
        "Analyst review / document in inventory",
        "N/A (informational)",
    ),
    (
        "transactions.csv",
        "outcome, close_date — A Closed deal must have completed, so a close date is required.",
        "Business rule DQ: Outcome Closed requires close_date",
        "No violations in sample (0 rows)",
        "Asset check fails if violated; does not auto-correct staging",
        "lower(trim(outcome)) = 'closed' AND close_date IS NULL → count must be 0",
        "william_blair_de/assets/checks.py (stg_transactions_closed_requires_close_date)",
    ),
    (
        "transactions.csv",
        "announce_date, close_date, days_to_close — Business days from announce to close for completed deals.",
        "Business rule DQ: both dates present requires days_to_close",
        "No violations in sample (0 rows)",
        "Asset check fails if violated; does not auto-correct staging",
        "announce_date AND close_date NOT NULL AND days_to_close IS NULL → count must be 0",
        "william_blair_de/assets/checks.py (stg_transactions_dates_require_days_to_close)",
    ),
    (
        "transactions.csv",
        "ev_ebitda_multiple, ev_revenue_multiple, synergy_pct_of_deal, num_bidders — Valuation richness, synergy, process competitiveness.",
        "(Not yet implemented) Extreme / impossible values",
        "Not coded as dedicated checks",
        "Would extend: quarantine or AssetCheckResult(severity=WARN) with thresholds",
        "Future: e.g. ev_ebitda_multiple BETWEEN 0 AND 100",
        "Add to william_blair_de/assets/checks.py",
    ),
]

INGEST_HEADERS = [
    "File name",
    "Field(s) — business meaning",
    "Issue",
    "Observed?",
    "How resolved",
    "Programmatic check",
    "Code location",
]

INGEST_ROWS = [
    (
        "All five CSVs",
        "All columns as loaded from CSV — Preserve upstream spelling, nulls, and formatting for audit and reprocessing.",
        "Accidental type coercion in ingest",
        "N/A",
        "Load with read_csv_auto(..., all_varchar=true)",
        "Ingest uses all-varchar",
        "william_blair_de/assets/raw.py (_load_csv_as_varchar)",
    ),
]

SQL_SNIPPET = """-- After materialization (DuckDB)
SELECT COUNT(*) FROM staging.transactions_quarantine;
SELECT * FROM staging.transactions_quarantine LIMIT 20;

-- Business rule DQ (same logic as asset checks)
SELECT COUNT(*) AS closed_without_close_date
FROM staging.transactions
WHERE lower(trim(coalesce(outcome, ''))) = 'closed' AND close_date IS NULL;

SELECT COUNT(*) AS both_dates_but_null_days_to_close
FROM staging.transactions
WHERE announce_date IS NOT NULL AND close_date IS NOT NULL AND days_to_close IS NULL;
"""


def _autosize(ws, max_width: int = 70) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 12
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            width = max(width, min(max_width, len(str(v)) + 2))
        ws.column_dimensions[letter].width = width


def main() -> None:
    wb = Workbook()

    # --- Sheet 1: inventory ---
    ws = wb.active
    ws.title = "DQ inventory"
    ws["A1"] = INTRO
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(HEADERS))
    ws.row_dimensions[1].height = 36

    start = 4
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=start, column=j, value=h)
        c.font = Font(bold=True)
    for i, row in enumerate(ROWS, start=start + 1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws)

    # --- Sheet 2: raw ingest ---
    w2 = wb.create_sheet("Raw ingest fidelity")
    for j, h in enumerate(INGEST_HEADERS, start=1):
        w2.cell(row=1, column=j, value=h).font = Font(bold=True)
    for i, row in enumerate(INGEST_ROWS, start=2):
        for j, val in enumerate(row, start=1):
            w2.cell(row=i, column=j, value=val).alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(w2)

    # --- Sheet 3: SQL ---
    w3 = wb.create_sheet("Post-materialization SQL")
    w3["A1"] = SQL_SNIPPET.strip()
    w3["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    w3.column_dimensions["A"].width = 90
    w3.row_dimensions[1].height = 60

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
